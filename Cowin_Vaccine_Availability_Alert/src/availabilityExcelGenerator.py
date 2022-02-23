from typing import List
import openpyxl
import datetime as dt

def generateAvailabilityExcel(finalOpList: List[dict], excelDumpPath:str, startDate:dt.datetime, endDate:dt.datetime):

    startDateStr = dt.datetime.strftime(startDate, '%Y-%m-%d')
    endDateStr = dt.datetime.strftime(endDate, '%Y-%m-%d')
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    currRowNo =2

    #setting column name
    sheet.cell(column=1, row=1, value= "Date")
    sheet.cell(column=2, row=1, value= "District_name")
    sheet.cell(column=3, row=1, value= "block_name")
    sheet.cell(column=4, row=1, value= "Address")
    sheet.cell(column=5, row=1, value= "Fee_type")
    sheet.cell(column=6, row=1, value= "Available_capacity")
    sheet.cell(column=7, row=1, value= "Available_capacity_dose1")
    sheet.cell(column=8, row=1, value= "Available_capacity_dose2")
    sheet.cell(column=9, row=1, value= "Vaccine_Name")
    sheet.cell(column=10, row=1, value= "Min_Age_Limit")
    sheet.cell(column=11, row=1, value= "Max_Age_Limit")

    for centre in finalOpList:
        sheet.cell(column=1, row=currRowNo, value= centre["date"])
        sheet.cell(column=2, row=currRowNo, value= centre["district_name"])
        sheet.cell(column=3, row=currRowNo, value= centre["block_name"])
        sheet.cell(column=4, row=currRowNo, value= centre["address"])
        sheet.cell(column=5, row=currRowNo, value= centre["fee_type"])
        sheet.cell(column=6, row=currRowNo, value= centre["available_capacity"])
        sheet.cell(column=7, row=currRowNo, value= centre["available_capacity_dose1"])
        sheet.cell(column=8, row=currRowNo, value= centre["available_capacity_dose2"])
        sheet.cell(column=9, row=currRowNo, value= centre["vaccine_name"])
        sheet.cell(column=10, row=currRowNo, value= centre["min_age_limit"])
        sheet.cell(column=11, row=currRowNo, value= centre["max_age_limit"])
        currRowNo = currRowNo+1

    # saving workbook with file name ex. 2022-02-23_to_2022-02-25.xlsx
    genFilename= f"{startDateStr}_to_{endDateStr}.xlsx "
    wb.save(filename=f"{excelDumpPath}/{genFilename}" )